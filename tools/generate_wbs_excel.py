from pathlib import Path
import os
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).parent))
from validate_wbs import ROOT, TASKS, validate

OUT = ROOT / "reports/wbs.xlsx"
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", MAIN_NS)

def extract_section(text, heading):
    lines = text.splitlines(); start = None
    for i, line in enumerate(lines):
        if line.strip() == f"# {heading}": start = i + 1; break
    if start is None: return "未登録"
    end = len(lines)
    for i in range(start, len(lines)):
        if lines[i].startswith("# "): end = i; break
    value = "\n".join(lines[start:end]).strip()
    return value or "未登録"

def _value(meta, key):
    value = meta.get(key)
    return ", ".join(value) if isinstance(value, list) else (value or "")

def _set_read_only_recommended(source, output):
    source = Path(source); output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    fd, patched_name = tempfile.mkstemp(prefix="wbs-patched-", suffix=".xlsx", dir=output.parent)
    os.close(fd)
    try:
        with zipfile.ZipFile(source, "r") as source_zip, zipfile.ZipFile(patched_name, "w", zipfile.ZIP_DEFLATED) as target_zip:
            workbook_xml = ET.fromstring(source_zip.read("xl/workbook.xml"))
            sharing_nodes = workbook_xml.findall(f"{{{MAIN_NS}}}fileSharing")
            sharing = sharing_nodes[0] if sharing_nodes else ET.Element(f"{{{MAIN_NS}}}fileSharing")
            for duplicate in sharing_nodes:
                workbook_xml.remove(duplicate)
            sharing.set("readOnlyRecommended", "1")
            children = list(workbook_xml)
            anchors = ("workbookPr", "bookViews", "sheets", "definedNames", "calcPr", "extLst")
            insert_at = next((i for i, child in enumerate(children) if child.tag.rsplit("}", 1)[-1] in anchors), len(children))
            workbook_xml.insert(insert_at, sharing)
            workbook_data = ET.tostring(workbook_xml, encoding="utf-8", xml_declaration=True)
            for info in source_zip.infolist():
                target_zip.writestr(info, workbook_data if info.filename == "xl/workbook.xml" else source_zip.read(info.filename))
        os.replace(patched_name, output)
    except Exception:
        try: os.unlink(patched_name)
        except FileNotFoundError: pass
        raise

def _tree(data):
    children = {ident: [] for ident in data}
    for ident, (_, meta, _) in data.items():
        if meta.get("parent") in children: children[meta["parent"]].append(ident)
    order = []
    def visit(ident, level):
        order.append((ident, level))
        for child in children[ident]: visit(child, level + 1)
    for ident, (_, meta, _) in data.items():
        if not meta.get("parent"): visit(ident, 0)
    return order, children

def generate_wbs_excel(tasks_dir=TASKS, output=OUT):
    data, issues = validate(tasks_dir, Path(tasks_dir).resolve().parents[1])
    if any(x[0] == "error" for x in issues): raise ValueError("WBS validation failed; Excel was not generated")
    order, children = _tree(data); wb = Workbook(); ws = wb.active; ws.title = "WBS一覧"
    headers = ["WBS ID", "プロジェクト番号", "プロジェクト内番号", "親WBS ID", "階層", "種別", "WBS名", "担当", "状態", "依存先", "流用元WBS", "仕様書", "関連ファイル", "成果物", "ブロック理由", "完了条件要約", "証跡要約", "更新日", "詳細ファイル"]
    ws.append(headers)
    for ident, level in order:
        path, meta, text = data[ident]; task = meta.get("type") == "task"
        try: detail_path = str(path.relative_to(ROOT))
        except ValueError: detail_path = str(path.relative_to(Path(tasks_dir).resolve().parent))
        ws.append([ident, ident[4:7], ident[8:], meta.get("parent") or "", level, meta.get("type"), meta.get("title"), meta.get("actor"), meta.get("status"), _value(meta, "depends_on"), _value(meta, "source_wbs"), _value(meta, "specs"), _value(meta, "related_files"), _value(meta, "outputs"), _value(meta, "blocked_by"), extract_section(text, "完了条件") if task else "", extract_section(text, "証跡") if task else "", meta.get("updated"), detail_path])
    ws2 = wb.create_sheet("未完了一覧"); ws2.append(["WBS ID", "WBS名", "担当", "親WBS ID", "依存先", "ブロック理由", "完了条件要約", "詳細ファイル"])
    for ident, (_, meta, text) in data.items():
        if meta.get("type") == "task" and meta.get("status") == "incomplete": ws2.append([ident, meta.get("title"), meta.get("actor"), meta.get("parent") or "", _value(meta, "depends_on"), _value(meta, "blocked_by"), extract_section(text, "完了条件"), f"wbs/tasks/{ident}.md"])
    tasks = [m for _, m, _ in data.values() if m.get("type") == "task"]; done = [m for m in tasks if m.get("status") == "complete"]
    ws3 = wb.create_sheet("進捗集計"); ws3.append(["項目", "値"])
    stats = [("実行タスク総数", len(tasks)), ("完了タスク数", len(done)), ("未完了タスク数", len(tasks) - len(done)), ("完了率", len(done) / len(tasks) if tasks else 0), ("Codex担当タスク数", sum(m.get("actor") == "codex" for m in tasks)), ("Codex担当の未完了数", sum(m.get("actor") == "codex" and m.get("status") != "complete" for m in tasks)), ("人担当タスク数", sum(m.get("actor") == "human" for m in tasks)), ("人担当の未完了数", sum(m.get("actor") == "human" and m.get("status") != "complete" for m in tasks)), ("担当未確定数", sum(m.get("actor_review_required", False) for m in tasks)), ("グループ数", len(data) - len(tasks))]
    for row in stats: ws3.append(row)
    ws3.append([]); ws3.append(["最上位group別集計"]); ws3.append(["WBS ID", "WBS名", "配下task総数", "完了task数", "未完了task数", "完了率", "Codex未完了", "人未完了", "担当要確認"])
    roots = [i for i, (_, m, _) in data.items() if m.get("type") == "group" and not m.get("parent")]
    for root in roots:
        descendants = []
        def collect(i):
            for child in children[i]:
                if data[child][1].get("type") == "task": descendants.append(data[child][1])
                collect(child)
        collect(root); completed = sum(m.get("status") == "complete" for m in descendants)
        ws3.append([root, data[root][1].get("title"), len(descendants), completed, len(descendants) - completed, completed / len(descendants) if descendants else 0, sum(m.get("actor") == "codex" and m.get("status") != "complete" for m in descendants), sum(m.get("actor") == "human" and m.get("status") != "complete" for m in descendants), sum(m.get("actor_review_required", False) for m in descendants)])
    root_tasks = [m for _, m, _ in data.values() if m.get("type") == "task" and not m.get("parent")]
    if root_tasks:
        completed = sum(m.get("status") == "complete" for m in root_tasks)
        ws3.append(["ROOT-TASKS", "ルートtask", len(root_tasks), completed, len(root_tasks) - completed, completed / len(root_tasks), sum(m.get("actor") == "codex" and m.get("status") != "complete" for m in root_tasks), sum(m.get("actor") == "human" and m.get("status") != "complete" for m in root_tasks), sum(m.get("actor_review_required", False) for m in root_tasks)])
    ws4 = wb.create_sheet("検証結果"); ws4.append(["種別", "WBS ID", "ファイル", "内容"])
    for issue in issues: ws4.append(issue)
    ws5 = wb.create_sheet("移行対応"); ws5.append(["旧階層パス（parent/titleから復元）", "新WBS ID", "種別", "担当", "状態", "分割・統合情報", "備考"])
    for ident, level in order:
        path = []; cur = ident
        while cur:
            path.append(data[cur][1].get("title")); cur = data[cur][1].get("parent")
        meta = data[ident][1]; ws5.append([" > ".join(reversed(path)), ident, meta.get("type"), meta.get("actor"), meta.get("status"), "なし", "PROJECT.md削除後のparent/titleから復元"])
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in sheet.iter_rows():
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
        for i in range(1, sheet.max_column + 1): sheet.column_dimensions[get_column_letter(i)].width = min(52, max(12, max(len(str(sheet.cell(r, i).value or "")) for r in range(1, min(sheet.max_row, 40) + 1)) + 2))
    output = Path(output); output.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix="wbs-generated-", suffix=".xlsx", dir=output.parent)
    os.close(fd)
    try:
        wb.save(temporary_name)
        _set_read_only_recommended(temporary_name, output)
    except Exception:
        try: os.unlink(temporary_name)
        except FileNotFoundError: pass
        raise
    finally:
        try: os.unlink(temporary_name)
        except FileNotFoundError: pass
    return output

def main():
    try: print(generate_wbs_excel())
    except ValueError as exc: print(exc, file=sys.stderr); return 1
    return 0

if __name__ == "__main__": sys.exit(main())
