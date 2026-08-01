import tempfile
import unittest
from pathlib import Path
from openpyxl import load_workbook
from tools.generate_wbs_excel import extract_section, generate_wbs_excel

class GenerateWbsExcelTest(unittest.TestCase):
    def test_section_extraction_and_empty_section(self):
        text = "# 完了条件\n\n条件1\n条件2\n# 証跡\n\n証跡1\n# その他\n\n除外"
        self.assertEqual("条件1\n条件2", extract_section(text, "完了条件")); self.assertEqual("証跡1", extract_section(text, "証跡")); self.assertEqual("未登録", extract_section("# 完了条件\n\n# 証跡\n\n", "完了条件"))

    def test_current_workbook_shape_counts_and_strings(self):
        book = load_workbook(Path(__file__).parents[1] / "reports/wbs.xlsx")
        self.assertEqual({"WBS一覧", "未完了一覧", "進捗集計", "検証結果", "移行対応"}, set(book.sheetnames))
        ws = book["WBS一覧"]; self.assertEqual(674, ws.max_row); self.assertEqual("WBS-001-01000", ws["A2"].value); self.assertEqual("001", ws["B2"].value); self.assertEqual("01000", ws["C2"].value); self.assertIsInstance(ws["A2"].value, str)
        self.assertEqual(143, book["未完了一覧"].max_row)
        summary = [row[0].value for row in book["進捗集計"].iter_rows()]; self.assertIn("最上位group別集計", summary)
        self.assertTrue(any("既存項目を完了し証跡を記録する" not in str(cell.value) for row in ws.iter_rows(min_row=2, max_row=3) for cell in row))
        self.assertIn(" > ", book["移行対応"]["A3"].value); self.assertIn("PROJECT.md削除後", book["移行対応"]["G2"].value)

    def test_small_wbs_can_be_generated_to_temporary_output(self):
        with tempfile.TemporaryDirectory() as d:
            tasks = Path(d) / "tasks"; tasks.mkdir(); p = tasks / "WBS-001-01000.md"
            p.write_text('''---\nid: WBS-001-01000\ntitle: 小規模\ntype: task\nstatus: incomplete\nactor: codex\nparent: null\ndepends_on: []\nspecs: []\nrelated_files: []\noutputs: []\nblocked_by: []\nsource_wbs: []\nevidence_required: true\nupdated: '2026-08-01'\n---\n# 目的\n\n目的\n# 作業内容\n\n作業\n# 入力\n\n入力\n# 成果物\n\n成果物\n# 完了条件\n\n小規模条件\n# 完了にしてはいけない条件\n\n禁止\n# 確認結果\n\n確認\n# 証跡\n\n小規模証跡\n''', encoding="utf-8")
            output = Path(d) / "out.xlsx"; generate_wbs_excel(tasks, output); ws = load_workbook(output)["WBS一覧"]
            self.assertEqual("小規模条件", ws["P2"].value); self.assertEqual("小規模証跡", ws["Q2"].value)

if __name__ == "__main__": unittest.main()
